"""
Test Case Generator - XRecorder 专用脚本
列：Unnamed: 0（模块）| 用例名称 | 操作 | 预期 | 优先级 | 备注

用法：
    python generate_xrecorder.py --input cases.csv --output output.xlsx --changes changes.json

changes.json 格式：
{
  "modified": [
    {
      "case": "2K/4K新功能弹窗",   ← 用例名称定位（推荐）
      "col": "C",                  ← C=操作 D=预期 E=优先级 F=备注
      "runs": [
        {"text": "原有内容\\n", "red": false},
        {"text": "新增内容\\n", "red": true}
      ]
    }
  ],
  "new_rows": [
    {
      "after_case": "2K/4K新功能弹窗",   ← 插入到此用例名称的正下方
      "data": {
        "Unnamed: 0": "",
        "用例名称": "「隐形悬浮球」新功能弹窗",
        "操作": "1.xxx\\n2.xxx",
        "预期": "1.xxx\\n2.xxx",
        "优先级": "",
        "备注": ""
      }
    }
  ],
  "deprecated": ["用例名称A", "用例名称B"]   ← 用例名称列表，备注列追加「已废弃」
}

注意：
- COLUMNS 顺序：Unnamed: 0 | 用例名称 | 操作 | 预期 | 优先级 | 备注
- 列字母映射：A=模块 B=用例名称 C=操作 D=预期 E=优先级 F=备注
- modified 用 case（用例名称）定位，不受新行插入影响
- new_rows 用 after_case 精确定位插入位置；after_case 为空则追加到末尾
"""

import argparse
import json
import os
import shutil
import zipfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── 常量 ────────────────────────────────────────────────────

COLUMNS = ['Unnamed: 0', '用例名称', '操作', '预期', '优先级', '备注']
HEADERS = ['模块', '用例名称', '操作', '预期', '优先级', '备注']

# 列字母映射（A-F）
COL_LETTERS = {col: get_column_letter(i + 1) for i, col in enumerate(COLUMNS)}
LETTER_TO_COL = {v: k for k, v in COL_LETTERS.items()}

COL_WIDTHS = {'A': 18, 'B': 24, 'C': 60, 'D': 60, 'E': 8, 'F': 30}

HEADER_COLOR = '2F4F4F'
RED_COLOR = 'EA4335'
DEPRECATED_NOTE = '已废弃'
NEW_ROW_MARKER = '__is_new__'


# ── 富文本修复 ──────────────────────────────────────────────

def fix_rich_text_xlsx(filepath):
    tmp = filepath + '.tmp'
    shutil.copy(filepath, tmp)
    with zipfile.ZipFile(tmp, 'r') as z:
        sheet_bytes = z.read('xl/worksheets/sheet1.xml')
        all_files = {n: z.read(n) for n in z.namelist()}
    fixed = sheet_bytes.replace(b'rgb="00000000"', b'rgb="FF000000"')
    fixed = fixed.replace(b'rgb="00EA4335"', b'rgb="FFEA4335"')
    fixed = fixed.replace(b'<sz val="1000"/>', b'<sz val="10"/>')
    all_files['xl/worksheets/sheet1.xml'] = fixed
    with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in all_files.items():
            zout.writestr(name, data)
    os.remove(tmp)


# ── 富文本写入 ──────────────────────────────────────────────

def make_rich_cell(cell, runs):
    blocks = []
    for run in runs:
        color = RED_COLOR if run['red'] else '000000'
        ifont = InlineFont(rFont='Arial', sz=1000, color=color)
        blocks.append(TextBlock(ifont, run['text']))
    cell.value = CellRichText(*blocks)
    cell.alignment = Alignment(wrap_text=True, vertical='top')


# ── 样式工具 ────────────────────────────────────────────────

def style_header(ws):
    fill = PatternFill('solid', start_color=HEADER_COLOR)
    font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_data_cell(cell, red=False):
    cell.font = Font(name='Arial', size=10, color=RED_COLOR if red else '000000')
    cell.alignment = Alignment(wrap_text=True, vertical='top')


# ── CSV 加载 ────────────────────────────────────────────────

def load_csv(path):
    df = pd.read_csv(path, dtype=str, encoding='utf-8').fillna('')
    missing = [c for c in COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"CSV 缺少列：{missing}，实际列名：{df.columns.tolist()}")
    return df[COLUMNS].copy()


# ── 新行插入（按 after_case 精确定位）──────────────────────

def insert_new_rows(df, new_rows):
    df = df.copy()
    df[NEW_ROW_MARKER] = False

    for item in new_rows:
        after_case = item.get('after_case', '').strip()
        row_data = item['data']

        if after_case:
            # 找到 after_case 对应的最后一个匹配行（支持同名用例多行）
            matches = df[df['用例名称'].str.strip() == after_case]
            if matches.empty:
                print(f"⚠️  警告：找不到 after_case='{after_case}'，追加到末尾")
                insert_pos = len(df)
            else:
                insert_pos = matches.index[-1] + 1
        else:
            insert_pos = len(df)

        new_row = {c: row_data.get(c, '') for c in COLUMNS}
        new_row[NEW_ROW_MARKER] = True
        new_df = pd.DataFrame([new_row])
        df = pd.concat([df.iloc[:insert_pos], new_df, df.iloc[insert_pos:]], ignore_index=True)
        df = df.reset_index(drop=True)

    return df


# ── modified 解析（按 case 定位）───────────────────────────

def resolve_modified_map(df, modified_list):
    """
    将 modified 列表解析为 {(excel_row, col_letter): runs} 映射。

    XRecorder CSV 无独立模块列（模块合并单元格），用「用例名称」唯一定位。
    insert_new_rows 完成后调用，行号已稳定。
    """
    # 建立 case -> excel_row 查找表（用例名称应唯一，取第一个匹配）
    lookup = {}
    for df_idx, row in df.iterrows():
        excel_row = df_idx + 2  # +1 表头, +1 0-based
        case = str(row['用例名称']).strip()
        if case and case not in lookup:
            lookup[case] = excel_row

    modified_map = {}
    for mod in modified_list:
        col = mod.get('col', '')
        runs = mod.get('runs', [])

        if 'case' in mod:
            case_key = mod['case'].strip()
            if case_key not in lookup:
                print(f"⚠️  警告：找不到 case='{case_key}'，跳过")
                continue
            excel_row = lookup[case_key]
        elif 'row' in mod:
            # 兼容旧格式直接指定行号
            excel_row = mod['row']
        else:
            print(f"⚠️  警告：modified 条目缺少 case 或 row 字段，跳过")
            continue

        modified_map[(excel_row, col)] = runs

    return modified_map


# ── 主构建函数 ──────────────────────────────────────────────

def build_xlsx(df, changes, output_path):
    deprecated_cases = set(changes.get('deprecated', []))
    df = insert_new_rows(df, changes.get('new_rows', []))
    modified_map = resolve_modified_map(df, changes.get('modified', []))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.freeze_panes = 'A2'

    for letter, width in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    # 表头
    ws.append(HEADERS)
    style_header(ws)
    ws.row_dimensions[1].height = 20

    for df_idx, row in df.iterrows():
        excel_row = df_idx + 2
        is_new = bool(row.get(NEW_ROW_MARKER, False))
        case_name = str(row['用例名称']).strip()
        is_deprecated = case_name in deprecated_cases

        ws.append([row[c] for c in COLUMNS])

        # 估算行高（按操作/预期列换行数）
        max_lines = 1
        for col in ['操作', '预期']:
            val = str(row[col])
            max_lines = max(max_lines, val.count('\n') + 1)
        ws.row_dimensions[excel_row].height = max(20, max_lines * 15)

        for col_idx, col_letter in enumerate('ABCDEF', start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            key = (excel_row, col_letter)

            if key in modified_map:
                make_rich_cell(cell, modified_map[key])
            elif is_new:
                style_data_cell(cell, red=True)
            else:
                # 废弃：备注列追加「已废弃」
                if is_deprecated and col_letter == 'F':
                    orig_val = str(cell.value or '').strip()
                    cell.value = (orig_val + ' ' + DEPRECATED_NOTE).strip()
                style_data_cell(cell, red=False)

    wb.save(output_path)
    fix_rich_text_xlsx(output_path)
    print(f"✅ 已生成：{output_path}")


# ── 入口 ────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='XRecorder 测试用例 xlsx 生成器')
    parser.add_argument('--input',   required=True,  help='输入 CSV 路径')
    parser.add_argument('--output',  required=True,  help='输出 xlsx 路径')
    parser.add_argument('--changes', required=True,  help='changes.json 路径')
    args = parser.parse_args()

    df = load_csv(args.input)
    with open(args.changes, encoding='utf-8') as f:
        changes = json.load(f)

    build_xlsx(df, changes, args.output)


if __name__ == '__main__':
    main()
