"""
Test Case Generator - 标准五列格式专用脚本
列：模块 | 用例名称 | 描述 | 预期 | 备注

用法：
    python generate.py --input cases.csv --output output.xlsx --changes changes.json

changes.json 格式：
{
  "modified": [
    {
      "module": "撤销/重做",
      "case": "重做（Redo）",
      "col": "C",
      "runs": [...]
    }
  ],
  ...
}

注意：
- modified 使用 module+case 定位行，不受新行插入影响
- case 字段始终填写 CSV 原始用例名称（即使 B 列同时有改名条目）
- 兼容旧版 row 字段（若无 module/case 则回退到 row 定位）
"""

import argparse
import json
import os
import shutil
import zipfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill


COLUMNS = ['模块', '用例名称', '描述', '预期', '备注']
COL_WIDTHS = {'A': 18, 'B': 16, 'C': 50, 'D': 45, 'E': 14}
HEADER_COLOR = '2F4F4F'
RED_COLOR = 'EA4335'
DEPRECATED_NOTE = '已废弃'
NEW_ROW_MARKER = '__is_new__'


# ── 富文本修复 ──────────────────────────────────────────────

def fix_rich_text_xlsx(filepath):
    tmp = filepath + '.tmp'
    shutil.copy(filepath, tmp)
    with zipfile.ZipFile(tmp, 'r') as z:
        sheet_bytes = z.read('xl/worksheets/sheet1.xml')
        all_files = {n: z.read(n) for n in z.namelist()}
    fixed = sheet_bytes.replace(b'rgb="00000000"', b'rgb="FF000000"')
    fixed = fixed.replace(b'rgb="00EA4335"', b'rgb="FFEA4335"')
    fixed = fixed.replace(b'<sz val="1000"/>', b'<sz val="10"/>')
    all_files['xl/worksheets/sheet1.xml'] = fixed
    with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in all_files.items():
            zout.writestr(name, data)
    os.remove(tmp)


# ── 富文本写入 ──────────────────────────────────────────────

def make_rich_cell(cell, runs):
    blocks = []
    for run in runs:
        color = RED_COLOR if run['red'] else '000000'
        ifont = InlineFont(rFont='Arial', sz=1000, color=color)
        blocks.append(TextBlock(ifont, run['text']))
    cell.value = CellRichText(*blocks)
    cell.alignment = Alignment(wrap_text=True, vertical='top')


# ── 样式工具 ────────────────────────────────────────────────

def style_header(ws):
    fill = PatternFill('solid', start_color=HEADER_COLOR)
    font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_data_cell(cell, red=False):
    cell.font = Font(name='Arial', size=10, color=RED_COLOR if red else '000000')
    cell.alignment = Alignment(wrap_text=True, vertical='top')


# ── 主流程 ──────────────────────────────────────────────────

def load_csv(path):
    df = pd.read_csv(path, dtype=str, encoding='utf-8').fillna('')
    missing = [c for c in COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"CSV 缺少列：{missing}")
    return df[COLUMNS].copy()


def insert_new_rows(df, new_rows):
    df = df.copy()
    df[NEW_ROW_MARKER] = False
    for item in new_rows:
        module = item['after_module']
        row_data = item['data']
        # 매번 df_filled를 재계산하여 이미 삽입된 새 행들의 모듈명도 반영
        df_filled = df.copy()
        # 빈 모듈명 ffill: 단, 새로 삽입된 행은 모듈명이 이미 채워져 있음
        filled_module = df_filled['模块'].copy()
        last = ''
        for i in df_filled.index:
            val = filled_module.iloc[i]
            if val and val != '':
                last = val
            else:
                filled_module.iloc[i] = last
        df_filled['模块'] = filled_module
        mask = df_filled['模块'] == module
        insert_pos = df_filled[mask].index[-1] + 1 if mask.any() else len(df)
        new_row = {c: row_data.get(c, '') for c in COLUMNS}
        new_row[NEW_ROW_MARKER] = True
        new_df = pd.DataFrame([new_row])
        df = pd.concat([df.iloc[:insert_pos], new_df, df.iloc[insert_pos:]], ignore_index=True)
        df = df.reset_index(drop=True)
    return df


def resolve_modified_map(df, modified_list):
    """
    将 modified 列表解析为 {(excel_row, col_letter): runs} 映射。

    支持两种定位方式：
      1. module + case（推荐）：插入新行后自动定位
      2. row（兼容旧格式）：直接用 Excel 行号

    特殊处理：当同一用例同时有 col=B（改名）和其他列的条目时，
    col=B 的改名不影响其他列用原始 case 名查找——
    查找表同时记录原始 case 和改名后的 case，两者都能命中同一行。
    """
    df_filled = df.copy()
    df_filled['模块'] = df_filled['模块'].replace('', None).ffill()

    # 先收集所有 B 列改名：(module, old_case) -> new_case
    rename_map = {}
    for mod in modified_list:
        if mod.get('col') == 'B' and 'module' in mod and 'case' in mod:
            new_name = ''.join(r['text'] for r in mod['runs'])
            rename_map[(mod['module'].strip(), mod['case'].strip())] = new_name

    # 建立 lookup：(module_filled, case) -> excel_row
    # 同时为改名的条目登记新旧两个 case 名
    lookup = {}
    for df_idx, row in df_filled.iterrows():
        excel_row = df_idx + 2
        mod = str(row['模块']).strip()
        case = str(row['用例名称']).strip()
        lookup[(mod, case)] = excel_row
        # 如果这行被改名，也用新名注册
        if (mod, case) in rename_map:
            new_case = rename_map[(mod, case)]
            lookup[(mod, new_case)] = excel_row

    modified_map = {}
    for mod in modified_list:
        col = mod['col']
        runs = mod['runs']

        if 'module' in mod and 'case' in mod:
            key_lookup = (mod['module'].strip(), mod['case'].strip())
            if key_lookup not in lookup:
                print(f"⚠️  警告：找不到 module={mod['module']} case={mod['case']}，跳过")
                continue
            excel_row = lookup[key_lookup]
        elif 'row' in mod:
            excel_row = mod['row']
        else:
            print(f"⚠️  警告：modified 条目缺少 module+case 或 row 字段，跳过")
            continue

        modified_map[(excel_row, col)] = runs

    return modified_map


def build_xlsx(df, changes, output_path):
    deprecated_rows = set(changes.get('deprecated', []))
    df = insert_new_rows(df, changes.get('new_rows', []))
    modified_map = resolve_modified_map(df, changes.get('modified', []))

    wb = Workbook()
    ws = wb.active

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.append(COLUMNS)
    style_header(ws)

    orig_csv_counter = 0

    for df_idx, row in df.iterrows():
        excel_row = df_idx + 2
        is_new = bool(row.get(NEW_ROW_MARKER, False))

        ws.append([row[c] for c in COLUMNS])

        for col_idx, col_letter in enumerate(['A', 'B', 'C', 'D', 'E'], start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            key = (excel_row, col_letter)

            if key in modified_map:
                make_rich_cell(cell, modified_map[key])
            elif is_new:
                style_data_cell(cell, red=True)
            else:
                if orig_csv_counter in deprecated_rows and col_letter == 'E':
                    orig_val = cell.value or ''
                    cell.value = (orig_val + ' ' + DEPRECATED_NOTE).strip()
                style_data_cell(cell, red=False)

        if not is_new:
            orig_csv_counter += 1

        ws.row_dimensions[excel_row].height = 60

    wb.save(output_path)
    fix_rich_text_xlsx(output_path)
    print(f"✅ 已生成：{output_path}")


# ── 入口 ────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', required=True)
    parser.add_argument('--output', required=True)
    parser.add_argument('--changes', required=True)
    args = parser.parse_args()

    df = load_csv(args.input)
    with open(args.changes, encoding='utf-8') as f:
        changes = json.load(f)

    build_xlsx(df, changes, args.output)


if __name__ == '__main__':
    main()
